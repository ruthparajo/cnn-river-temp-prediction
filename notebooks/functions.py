import cv2
import datetime
import json
import numpy as np
import os
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import geopandas as gpd
import matplotlib.pyplot as plt
import rasterio
from rasterio.features import geometry_mask, rasterize
from rasterio.transform import from_origin
import re
import shutil
import subprocess
from skimage.transform import resize
from shapely.geometry import LineString, box, Point
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import train_test_split
from sentinelhub import MimeType, CRS, BBox, SentinelHubRequest, SentinelHubDownloadClient, \
    DataCollection, bbox_to_dimensions, DownloadRequest, SHConfig
from rasterio.windows import Window


from models import *


def first_day(month, year):
    return datetime.date(year, month, 1)

def last_day(any_day):
    # The day 28 exists in every month. 4 days later, it's always next month
    next_month = any_day.replace(day=28) + datetime.timedelta(days=4)
    # subtracting the number of the current day brings us back one month
    return next_month - datetime.timedelta(days=next_month.day)

def find_closest_point(points_gdf, quadrant_bbox):
    lat_min, lon_min, lat_max, lon_max = quadrant_bbox
    
    # Calcular el centro del cuadrante
    center_x = (lat_min + lat_max) / 2
    center_y = (lon_min + lon_max) / 2
    center_point = Point(center_y, center_x)
    # Calculate distances from the center of the quadrant to all points
    distances = points_gdf.geometry.apply(lambda p: p.distance(center_point))
    # Find the minimum distance
    min_distance = distances.min()
    # Filter points that match the minimum distance
    closest_points = points_gdf[distances == min_distance]
    return closest_points

def rasterize_linestrings(lines, transform, out_shape):
    # Convertir a geometría de raster
    shapes = ((geom, 1) for geom in lines)
    mask = rasterize(shapes, out_shape=out_shape, transform=transform)
    return mask

def get_data_request(time_interval, evalscript, bbox, size, data_type, folder):
    if data_type == 'lst':
        responses = [SentinelHubRequest.output_response('default', MimeType.TIFF)] 
    elif data_type == 'ndvi':
        responses = [
            SentinelHubRequest.output_response('default', MimeType.TIFF),
            SentinelHubRequest.output_response('ndvi_image', MimeType.PNG)]
        
    return SentinelHubRequest(
        evalscript=evalscript,
        input_data=[
            SentinelHubRequest.input_data(
                data_collection=DataCollection.LANDSAT_OT_L1,
                time_interval=time_interval,
                mosaicking_order='leastCC',
                maxcc=0.1
            )
        ],
        responses=responses,
        bbox=bbox,
        size=size,
        data_folder=folder,
        config=config
    )

def adjust_size(size, max_size=2500):
    """ Adjusts the size so that none of the axes exceed max_size, maintaining the aspect ratio """
    width, height = size
    max_dimension = max(width, height)

    if max_dimension > max_size:
        scale_factor = max_size / max_dimension
        width = int(width * scale_factor)
        height = int(height * scale_factor)

    return width, height

def get_data(shp, evalscript, time_intervals, data_type, folder):
    coords_wgs84 = list(shp.total_bounds)#[lon_min, lat_min, lon_max, lat_max]
    if data_type == 'lst':  # Assuming 'lst' uses thermal bands B10 or B11
        resolution = 100  # Set to 100 meters for thermal bands
    else:
        resolution = 30  # Default to 30 meters for other bands

    #resolution = 30
    bbox = BBox(bbox=coords_wgs84, crs=CRS.WGS84)
    # extract the size based on bbx and the resolution
    size = bbox_to_dimensions(bbox, resolution=resolution)

    size = adjust_size(size)

    # create a list of requests
    list_of_requests = [get_data_request(slot, evalscript, bbox, size, data_type, folder) for slot in time_intervals]
    list_of_requests = [request.download_list[0] for request in list_of_requests]

    # download data with multiple threads
    data = SentinelHubDownloadClient(config=config).download(list_of_requests, max_threads=5)


def load_raster(filepath,rgb = True):
  try:
      with rasterio.open(filepath) as src:
        if rgb:
          red = src.read(1)
          green = src.read(2)
          blue = src.read(3)
          rgb = np.dstack((red, green, blue))
          return rgb, src
        else:
          return src.read(1), src
  except rasterio.errors.RasterioIOError:
    print(f"{filepath} no es pot obrir com a TIFF.")
      
def resize_image(image, new_width, new_height):
    return resize(image, (new_height, new_width), preserve_range=True)

def extract_year_month_from_filename(filename):
    # Function to extract the year and month from the filename
    # Adjust regex to match the YYYY-MM format
    match = re.search(r'(\d{4})-(\d{2})', filename)
    if match:
        return match.group(0)  # Return the matched year and month as a string (YYYY-MM)
    return None

def load_data(dir_paths, W=None, rgb=[], show=None):
    data = {}
    time_slots = []
    for i, path in enumerate(dir_paths):
        image_files = os.listdir(path)

        # Filter out files without a valid date in the filename
        valid_image_files = [img for img in image_files if extract_year_month_from_filename(img) is not None]

        # Sort the valid image files by year and month extracted from the filename
        valid_image_files = sorted(valid_image_files, key=lambda x: extract_year_month_from_filename(x))
        
        times = []
        imgs_data = []
        for img_file in valid_image_files:
            imagen_path = os.path.join(path, img_file)
            try:
                img_data, meta = load_raster(imagen_path, rgb[i])  # Load raster data
    
                if W:  # Resize the image if W is provided
                    img_data = resize_image(img_data, W, W)
    
                imgs_data.append(img_data)
                times.append(extract_year_month_from_filename(img_file))
            except:
                pass
            
        data[path] = np.array(imgs_data)
        time_slots.append(times)

    return data, time_slots

def split_data(X, y, validation_split=0.1, test_split=0.1):
    """
    Splits the data into training, validation, and test sets.

    Parameters:
    X (np.array): Input data (e.g., images).
    y (np.array): Target data (e.g., temperature values).
    validation_split (float): Proportion of data to use for validation set.
    test_split (float): Proportion of data to use for test set.

    Returns:
    train_input (np.array): Training data inputs normalized.
    train_target (np.array): Training data targets.
    validation_input (np.array): Validation data inputs normalized.
    validation_target (np.array): Validation data targets.
    test_input (np.array): Test data inputs normalized.
    test_target (np.array): Test data targets.
    """

    n = X.shape[0]  # Number of samples
    rem_index = np.arange(n)  # Array of all indices

    # Select validation indices
    validation_index = np.random.choice(rem_index, int(validation_split * n), replace=False)
    rem_index = list(set(rem_index) - set(validation_index))

    # Select test indices from remaining
    test_index = np.random.choice(rem_index, int(test_split * n), replace=False)
    train_index = list(set(rem_index) - set(test_index))  # Remaining for training

    return train_index, validation_index, test_index#train_input, train_target, validation_input, validation_target, test_input, test_target

def evaluate_model(y_true, y_pred):
    """
    Returns useful metrics for evaluating a linear regression model for temperature prediction.

    Parameters:
    y_true: array-like, true temperature values.
    y_pred: array-like, predicted temperature values.

    Returns:
    dict containing all useful metrics.
    """

    # Mean Absolute Error (MAE)
    mae = mean_absolute_error(y_true, y_pred)

    # Mean Squared Error (MSE)
    mse = mean_squared_error(y_true, y_pred)

    # Root Mean Squared Error (RMSE)
    rmse = np.sqrt(mse)

    # R-squared (R²)
    r2 = r2_score(y_true, y_pred)

    # Mean Absolute Percentage Error (MAPE)
    epsilon = 1e-10  # Small value to avoid division by zero
    mape = np.mean(np.abs((y_true - y_pred) / (y_true + epsilon))) * 100

    # Mean Squared Error (MSE) sample-wise
    sample_wise_mse = []
    # Loop through each sample and compute the MSE for that sample
    for i in range(y_true.shape[0]):
        # Flatten the true and predicted values for this sample
        y_true_flatten = y_true[i].flatten()
        y_pred_flatten = y_pred[i].flatten()
        # Calculate MSE for this sample
        mse = mean_squared_error(y_true_flatten, y_pred_flatten)
        sample_wise_mse.append(mse)
    
    average_mse = np.mean(sample_wise_mse)

    # Create a dictionary with the metrics
    metrics = {
        'MAE': mae,
        'MSE': mse,
        'RMSE': rmse,
        'R²': r2,
        'MAPE (%)': mape,
        'MSE sample-wise': average_mse
    }

    return metrics

def update_excel_with_results(file_path, model_name, metrics, details):
    # Check if the Excel file exists
    if not os.path.exists(file_path):
        # Create a new Excel workbook and add a sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Model Results"
        # Write the header based on the metric keys
        header = ["Model Name"] + list(metrics.keys()) + ["Details"]
        sheet.append(header)
    else:
        # Load existing Excel workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active

    # Find the next empty row
    next_row = sheet.max_row + 1
    
    # Prepare the row with model name, metrics, and details
    row = [model_name] + list(metrics.values()) + [details]
    
    # Append the row data to the next available row
    for col_num, value in enumerate(row, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)
    
    # Save the Excel file
    workbook.save(file_path)

def save_excel(file_path, dicc, excel=None):
    # Check if the Excel file exists
    if not os.path.exists(file_path):
        # Create a new Excel workbook and add a sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        if excel == 'Results':
            sheet.title = "Model Results"
        else:
            sheet.title = "Model Details"
        # Write the header based on the metric keys
        header = list(dicc.keys())
        sheet.append(header)
    else:
        # Load existing Excel workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active

    # Find the next empty row
    next_row = sheet.max_row + 1
    
    # Prepare the row with model name, metrics, and details
    row = list(dicc.values())
    
    # Append the row data to the next available row
    for col_num, value in enumerate(row, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)
    
    # Save the Excel file
    workbook.save(file_path)

def split_data_df(X, y, validation_split=0.1, test_split=0.1):
    n = X.shape[0]  # Número de muestras
    rem_index = np.arange(n)  # Array de todos los índices

    # Seleccionar índices de validación
    validation_index = np.random.choice(rem_index, int(validation_split * n), replace=False)
    rem_index = list(set(rem_index) - set(validation_index))

    # Seleccionar índices de prueba de los restantes
    test_index = np.random.choice(rem_index, int(test_split * n), replace=False)
    train_index = list(set(rem_index) - set(test_index))  # El resto para entrenamiento

    # Dividir los datos en conjuntos de entrenamiento, validación y prueba
    validation_input = X.iloc[validation_index, :].values / 255.0  # Normalizar inputs si es necesario
    validation_target = y.iloc[validation_index].values

    test_input = X.iloc[test_index, :].values / 255.0  # Normalizar inputs
    test_target = y.iloc[test_index].values

    train_input = X.iloc[train_index, :].values / 255.0  # Normalizar inputs
    train_target = y.iloc[train_index].values

    return train_input, train_target, validation_input, validation_target, test_input, test_target

def plot_image(image, factor=1.0, clip_range=None, **kwargs):
    """
    Utility function for plotting RGB images.
    """
    fig, ax = plt.subplots(nrows=1, ncols=1, figsize=(15, 15))
    if clip_range is not None:
        ax.imshow(np.clip(image * factor, *clip_range), **kwargs)
    else:
        ax.imshow(image * factor, **kwargs)
    ax.set_xticks([])
    ax.set_yticks([])

def clear_directory(directory):
    # Function to clear the contents of a directory
    if os.path.exists(directory):
        for file in os.listdir(directory):
            file_path = os.path.join(directory, file)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.remove(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print(f'Error deleting {file_path}. Reason: {e}')

def save_raster(raster_array,filepath,shp):
    resolution = 30
    x_min, y_min, x_max, y_max = shp.total_bounds
    transform = from_origin(x_min, y_max, resolution, resolution)

    with rasterio.open(
        filepath,
        'w',
        driver='GTiff',
        height=raster_array.shape[0],
        width=raster_array.shape[1],
        count=len(raster_array.shape),
        dtype=raster_array.dtype,
        crs=shp.crs.to_string(),  # Ensure correct CRS
        transform=transform,
        nodata=0.0
    ) as dst:
        if len(raster_array.shape) == 3:
            dst.write(raster_array[:, :, 0], 1)  # Red channel
            dst.write(raster_array[:, :, 1], 2)  # Green channel
            dst.write(raster_array[:, :, 2], 3)  # Blue channel
        else:
            dst.write(raster_array, 1)

def distance_matrix(x0, y0, x1, y1):
    """
    Calculate distance matrix.
    Note: from <http://stackoverflow.com/questions/1871536>
    """

    obs = np.vstack((x0, y0)).T
    interp = np.vstack((x1, y1)).T

    d0 = np.subtract.outer(obs[:, 0], interp[:, 0])
    d1 = np.subtract.outer(obs[:, 1], interp[:, 1])

    # calculate hypotenuse
    return np.hypot(d0, d1)

def simple_idw(x, y, z, xi, yi, beta=2, dist_matrix=None):
    """
    Simple inverse distance weighted (IDW) interpolation
    x`, `y`,`z` = known data arrays containing coordinates and data used for interpolation
    `xi`, `yi` =  two arrays of grid coordinates
    `beta` = determines the degree to which the nearer point(s) are preferred over more distant points.
            Typically 1 or 2 (inverse or inverse squared relationship)
    """
    if dist_matrix is None:
        dist_matrix = distance_matrix(x, y, xi, yi)
    
    # Calculate weights using inverse distance with exponent `beta`
    weights = dist_matrix ** (-beta)
    weights /= weights.sum(axis=0) + 1e-12  # Normalizing weights

    weights = np.nan_to_num(weights, nan=0.0)
    z = np.nan_to_num(z, nan=0.0)

    return np.dot(weights.T, z)


    '''dist = distance_matrix(x, y, xi, yi)

    # In IDW, weights are 1 / distance
    # weights = 1.0/(dist+1e-12)**power
    weights = dist ** (-beta)
    
    # Make weights sum to one
    weights /= weights.sum(axis=0)

    weights = np.nan_to_num(weights, nan=0.0)
    z = np.nan_to_num(z, nan=0.0)

    # Multiply the weights for each interpolated point by all observed Z-values
    return np.dot(weights.T, z)'''

def project_linestrings_to_points(gdf,show=None):
    x_coords = []
    y_coords = []
    z_coords = []

    # Extract coordinates from LineStrings and store them as points
    for i, row in gdf.iterrows():
        geom = row.geometry
        if geom.geom_type == 'LineString':
            for coord in row.geometry.coords:
                x_coords.append(coord[0])
                y_coords.append(coord[1])
                if len(coord) == 3:
                    z_coords.append(coord[2])

        elif geom.geom_type == 'MultiLineString':
            # Si es MultiLineString, recorre cada LineString
            for linestring in geom.geoms:
                for coord in linestring.coords:
                    x_coords.append(coord[0])
                    y_coords.append(coord[1])
                    if len(coord) == 3:
                        z_coords.append(coord[2])

    if show:
        # Plot the points on a 2D grid
        plt.figure(figsize=(10, 10))
        plt.scatter(x_coords, y_coords, color='blue', s=10, label='Coordinates')
    
        # Customize the plot
        plt.title('Projected Coordinates (Points) in EPSG:2056', fontsize=15)
        plt.xlabel('Easting (meters)', fontsize=12)
        plt.ylabel('Northing (meters)', fontsize=12)
    
        # Add gridlines
        plt.grid(True)
    
        # Show the plot
        plt.legend()
        plt.show()
    return x_coords, y_coords, z_coords

def split_data_stratified(inputs, data_targets, labels):
    # 1. Obtener las etiquetas únicas
    unique_labels, label_indices = np.unique(labels, return_inverse=True)
    
    # 2. Hacer una división estratificada en función de las etiquetas únicas
    train_label_idx, temp_label_idx = train_test_split(unique_labels, test_size=0.4, random_state=42)
    
    # Dividir el conjunto temporal (val y test)
    val_label_idx, test_label_idx = train_test_split(temp_label_idx, test_size=0.5, random_state=42)
    
    # 3. Crear máscaras booleanas en base a las etiquetas
    train_indices = np.where(np.isin(labels, train_label_idx))[0]
    val_indices = np.where(np.isin(labels, val_label_idx))[0]
    test_indices = np.where(np.isin(labels, test_label_idx))[0]


    return train_indices, val_indices, test_indices

def load_river_raster(filepath):
    r, m = load_raster(filepath, False)
    name = filepath.split('/')[-1].split('.')[0].split('bw_')[-1]
    return name, r

def get_rivers_altitude(source_folder, filt):
    rivs=[]
    all_x =[]
    all_y=[]
    all_z=[]
    with rasterio.open('../../../../../data/simon.walther/swissAltitude/swissAltitude.vrt') as src:
        for f in os.listdir(source_folder):
            if os.path.join(source_folder, f).endswith('shp'):
                river = gpd.read_file(os.path.join(source_folder, f))
                x_coords, y_coords, z_coords = project_linestrings_to_points(river)
                
                if z_coords == []:
                    dir_altitudes = '../data/external/altitudes'
                    
                    for i in range(len(x_coords)):
                        lon, lat = x_coords[i], y_coords[i]
                        row, col = src.index(lon, lat)
                        window = Window(col_off=col, row_off=row, width=1, height=1)
                        pixel_value = src.read(1, window=window)
                        z_coords.append(pixel_value[0][0])
                       
                if filt and np.mean(z_coords) <= 800:
                    rivs.append(f.split('station_')[-1].split('.')[0])
                    all_x.append(x_coords)
                    all_y.append(y_coords)
                    all_z.append(z_coords)
                elif not filt:
                    rivs.append(f.split('station_')[-1].split('.')[0])
                    all_x.append(x_coords)
                    all_y.append(y_coords)
                    all_z.append(z_coords)
                    
                
    return rivs, all_x, all_y, all_z

# Function to compute Grad-CAM
def make_gradcam_heatmap(img_array, model, last_conv_layer_name, pred_index=None):
    """
    Generates a Grad-CAM heatmap.
    Args:
        img_array: Input image array (batch of shape (1, height, width, channels)).
        model: The model to analyze.
        last_conv_layer_name: The name of the last convolutional layer.
        pred_index: Index of the target class for Grad-CAM.
    Returns:
        A heatmap of the Grad-CAM.
    """
    print('holaaaa',img_array.shape)
    
    last_deep_layer = model.layers[-1].name
    
    grad_model = tf.keras.models.Model(
        inputs=model.inputs,
        outputs=[model.get_layer(last_conv_layer_name).output, model.get_layer(last_deep_layer).output]
    )
    print(grad_model.summary())

    # Record gradients
    with tf.GradientTape() as tape:
        last_conv_layer_output, preds = grad_model(img_array)
        if pred_index is None:
            pred_index = tf.argmax(preds[0])
        target_output = preds[:,0]
    print('Intento predir', target_output)
    # Compute the gradients
    grads = tape.gradient(target_output, last_conv_layer_output)
    print(grads)
    # Compute the mean intensity of the gradients
    pooled_grads = tf.reduce_mean(grads, axis=(0, 1, 2))

    # Multiply each channel in the feature map array by the pooled gradients
    last_conv_layer_output = last_conv_layer_output[0]
    heatmap = last_conv_layer_output @ pooled_grads[..., tf.newaxis]
    heatmap = tf.squeeze(heatmap)

    # Normalize the heatmap
    heatmap = tf.maximum(heatmap, 0) / tf.math.reduce_max(heatmap)
    return heatmap.numpy()

# Function to visualize Grad-CAM overlay
def display_grad_cam(img, heatmap, alpha=0.4):
    """
    Combines the original image with the generated heatmap.
    """
    heatmap = cv2.resize(heatmap, (img.shape[1], img.shape[0]))
    heatmap = np.uint8(255 * heatmap)
    heatmap = cv2.applyColorMap(heatmap, cv2.COLORMAP_JET)
    superimposed_img = cv2.addWeighted(img, 1 - alpha, heatmap, alpha, 0)
    plt.imshow(superimposed_img)
    plt.axis('off')
    plt.show()

def save_and_display_gradcam(img, heatmap, filename, alpha=0.4):
    """
    Saves and displays a Grad-CAM image.
    Args:
        img: The original image (H x W x C).
        heatmap: The heatmap generated by Grad-CAM.
        filename: File path to save the output image.
        alpha: Transparency factor for overlay.
    """
    # Resize heatmap to match input image
    heatmap = cv2.resize(heatmap, (img.shape[1], img.shape[0]))
    heatmap = np.uint8(255 * heatmap)
    heatmap = cv2.applyColorMap(heatmap, cv2.COLORMAP_JET)

    # Superimpose the heatmap on the original image
    superimposed_img = cv2.addWeighted(img, 1 - alpha, heatmap, alpha, 0)

    # Save the output image
    cv2.imwrite(filename, superimposed_img)
    print(f"Saved Grad-CAM to {filename}")

def save_grad_map(input_gradients, image_path, impact_channel, W, save_numpy=False):
    # Extraer gradientes para el canal de interés
    grad_map = input_gradients[0, :, :, impact_channel].numpy()  # (4, 4)

    # Opción para guardar los gradientes originales
    if save_numpy:
        np.save(f"{image_path}.npy", grad_map)  # Guardar los gradientes originales como .npy

    # Normalización para visualización
    grad_map_normalized = (grad_map - np.min(grad_map)) / (np.max(grad_map) - np.min(grad_map) + 1e-8)

    # Redimensionar para visualización
    grad_map_resized = tf.image.resize(grad_map_normalized[..., np.newaxis], (W, W)).numpy().squeeze()

    # Guardar la imagen para visualización
    plt.imshow(grad_map_resized, cmap="bwr")
    plt.colorbar()
    plt.title(f"Channel impact {impact_channel} in test")
    plt.savefig(f"{image_path}.png", dpi=300, bbox_inches='tight')
    plt.close()

    print('Saved at', image_path)


def load_all_data(
    source_folder='../data/external/shp/river_cells_oficial',
    source_path='../data/preprocessed/',
    data_paths=['lst', 'slope', 'discharge', 'ndvi', 'altitude'],
    filter_altitude=None,
    W=256,
    time_split=False,
):
    
    dir_paths = [os.path.join(source_path, p) for p in data_paths]
    
    all_dir_paths = {k: [] for k in data_paths}
    total_data = {}
    total_times = {}

    rivers = get_rivers_altitude(source_folder,filter_altitude)[0]
    
    # Cargar rutas de entrada
    for i, dir_p in enumerate(dir_paths):
        for subdir, dirs, files in os.walk(dir_p):
            if subdir != dir_p and not subdir.endswith('masked') and not subdir.endswith('.ipynb_checkpoints') and subdir.split('/')[-1] in rivers:
                all_dir_paths[data_paths[i]].append(subdir)
            elif dir_p.endswith('altitude'):
                all_dir_paths[data_paths[i]].extend([f for f in files if f.split('.')[0] in rivers])
      
    # Cargar datos de entrada
    
    for k, v in all_dir_paths.items():
        if k not in ['direction', 'slope', 'altitude']:
            labels = []
            list_rgb = [True] * len(v) if k in ['lst', 'masked'] else [False] * len(v)
            data, times = load_data(v, W, list_rgb)
            if k != 'masked':
                for ki in data.keys():
                    labels += [ki.split('/')[-1]] * len(data[ki])

            data_values = [np.array(img) for sublist in list(data.values()) for img in sublist]
            times_list = [t for sublist in times for t in sublist]

            if time_split:
                dates = [datetime.strptime(date, '%Y-%m') for date in times_list]
                pairs = sorted(zip(dates, data_values, labels), key=lambda x: x[0])
                sorted_dates, data_values, labels = zip(*pairs)
                times_list = [date.strftime('%Y-%m') for date in sorted_dates]

            total_data[k] = np.array(data_values)
            total_times[k] = times_list
            print(f"{k} : {total_data[k].shape}")

    origin_folder = source_path.split(f'{W}x{W}')[0]
    print(origin_folder,'ole')
    # Cargar variables adicionales
    for k, v in all_dir_paths.items():
        if k in ['direction', 'slope', 'altitude']:
            imgss = {}
            total = []
            for i, lab in enumerate(labels):
                for file in v:
                    if lab in file.split('/')[-1] or lab in file.split('.')[0]:
                        if lab not in imgss:
                            file_path = os.path.join(file, os.listdir(file)[0]) if k != 'altitude' else os.path.join(f'{source_path}altitude', file)
                            r, m = load_raster(file_path, False)
                            var = resize_image(r, W, W)
                            var = np.where(np.isnan(var), 0.0, var)
                            imgss[lab] = var
                        else:
                            var = imgss[lab]
                            
                total.append(var)

            total_data[k] = np.array(total)
            print(f"{k}: {np.array(total).shape}")

    # Cargar variable objetivo
    water_temp = pd.read_csv(f'{origin_folder}wt/water_temp.csv', index_col=0)
    times_ordered = total_times['lst']
    wt_temp = []
    for cell, date in zip(labels, times_ordered):
        temp = water_temp[(water_temp["Cell"] == cell) & (water_temp["Date"] == date)]["WaterTemp"]
        if not temp.empty:
            wt_temp.append(temp.values[0])
    data_targets = np.array(wt_temp)

    return total_data, total_times, data_targets, labels
    
def get_results(test_target, test_prediction, rivers, labels, test_index):
    mean_results = {k:[] for k in results.keys()}
    # Loop through each sample and compute the MSE for that sample
    for i in range(test_target.shape[0]):
        # Flatten the true and predicted values for this sample
        riv = rivers[labels[test_index[i]]].flatten()
        y_true_flatten = test_target[i].flatten()
        y_true_mask = y_true_flatten[riv != 0]
        y_pred_flatten = test_prediction[i].flatten()
        y_pred_mask = y_pred_flatten[riv != 0]
        # Calculate metrics
        res = evaluate_model(y_true_mask, y_pred_mask)
        for k,v in res.items():
            mean_results[k].append(v)
    for key in mean_results:
        mean_results[key] = np.mean(mean_results[key])
    return mean_results


def get_months_vectorized(times):
    """
    Calculate the cosine and sine values for each month in a given list of dates.
    Additionally, return a dictionary that maps each unique cosine value to its corresponding month.
    
    Parameters:
    ----------
    times : list or array-like
        A list of date strings or datetime objects from which month information is extracted.
        
    Returns:
    -------
    cosine_months : np.ndarray
        An array of cosine values corresponding to each month in the `times` input.
        
    sine_months : np.ndarray
        An array of sine values corresponding to each month in the `times` input.
        
    cos_to_month : dict
        A dictionary where each key is a unique cosine value and the corresponding value is the month (1-12) 
        associated with that cosine value."""
    
    month_cosine_dict = {month: np.cos((month - 1) / 12 * 2 * np.pi) for month in range(1, 13)}
    month_sinus_dict = {month: np.sin((month - 1) / 12 * 2 * np.pi) for month in range(1, 13)}
    
    cos_to_month = {cos_val: month for month, cos_val in month_cosine_dict.items()}
    
    def cosine_for_month(month):
        return month_cosine_dict[month]

    def sine_for_month(month):
        return month_sinus_dict[month]

    cosine_vectorized = np.vectorize(cosine_for_month)
    sine_vectorized = np.vectorize(sine_for_month)

    times_dt = pd.to_datetime(times)
    months = times_dt.month

    cosine_months = cosine_vectorized(months)
    sine_months = sine_vectorized(months)
    
    return cosine_months, sine_months, cos_to_month

def get_lat_lon(labels):
    file_path = '../data/raw/wt/cell_coordinates_oficial.csv'
    lat_lon = pd.read_csv(file_path)
    lats=[]
    lons=[]
    for label in labels:
        num_cell = int(label.split('_')[-1])
        lat = lat_lon[lat_lon.Cell==num_cell].Latitude.values[0]
        lon = lat_lon[lat_lon.Cell==num_cell].Longitude.values[0]
        lats.append(lat)
        lons.append(lon)
    lats = np.array(lats)
    lons = np.array(lons)
    return lats, lons

def get_split_index(split, input_data, data_targets, labels, split_id=None,filt_alt=None):
    if split == 'time':
        train_ratio = 0.6
        val_ratio = 0.2
        test_ratio = 0.2
        
        # Calcular el tamaño de cada conjunto
        total_images = len(input_data)
        train_size = int(total_images * train_ratio)
        val_size = int(total_images * val_ratio)
        indices = np.arange(total_images)
        
        train_index = indices[:train_size]                       # Primeros índices para entrenamiento
        validation_index = indices[train_size:train_size + val_size]    # Siguientes índices para validación
        test_index = indices[train_size + val_size:]             # Últimos índices para prueba
       
    elif split == 'stratified':
        print('yes !!!!!!!!!!')
        split_folder = f"../data/external/splits_low_alt/split_{split_id}_indices.json" if filt_alt else \
                        f"../data/external/splits/split_{split_id}_indices.json"
        with open(split_folder, 'r') as f:
            loaded_indices = json.load(f)
        train_index = loaded_indices['train']
        validation_index = loaded_indices['val']
        test_index = loaded_indices['test']
        #train_index, validation_index, test_index = split_data_stratified(input_data, data_targets, labels)
    else:
        train_index, validation_index, test_index = split_data(input_data, data_targets)
        
    return train_index, validation_index, test_index
            
def get_discharge(labels, times_ordered):
    discharge = pd.read_csv('../data/preprocessed/discharge/discharge.csv', index_col=0)
    disch = []
    for cell, date in zip(labels, times_ordered):
        value = discharge[(discharge["Cell"] == cell) & (discharge["Date"] == date)]["Discharge"]
        if not value.empty:
            disch.append(value.values[0])
    disch = np.array(disch)
    return disch

def get_next_experiment_number(file_path):
    if not os.path.exists(file_path):
        return 1 
    df = pd.read_excel(file_path)
    if df.empty:
        return 1
    return len(df) + 1

def normalize_min_max(data, min_val, max_val):
    """
    Normalizes the data to the range [0, 1] based on global min and max.
    
    Parameters:
    data (ndarray): Input data to normalize.
    min_val (float): Minimum value for normalization.
    max_val (float): Maximum value for normalization.
    
    Returns:
    ndarray: Normalized data.
    """
    if min_val == max_val:
        return np.zeros_like(data)  # Avoid division by zero
    return (data - min_val) / (max_val - min_val)

def build_model_map(model_name, input_args, W):
    additional = len(input_args)==2
    model_map = {
        "baseline_CNN": lambda: build_cnn_model_features(input_args[0], input_args[1]) if additional and W not in [8, 16] \
                                else (build_cnn_baseline_8x8(input_args) if W in [8, 16] and not additional else \
                                      (build_cnn_baseline(input_args) if not additional else \
                                       build_cnn_model_features_8x8(input_args[0], input_args[1]))
        ),
        "CNN_2": lambda: build_cnn_model_features2(input_args[0], input_args[1]) if additional and W not in [8, 16] \
                                else (build_cnn_baseline_8x8(input_args) if W in [8, 16] and not additional else \
                                      (build_cnn_baseline(input_args) if not additional else \
                                       build_cnn_model_features_8x8(input_args[0], input_args[1]))),
        "CNN_3": lambda: build_cnn_model_features3(input_args[0], input_args[1]) if additional else \
                            (build_cnn3(input_args)),
        "Resnet": lambda: build_resnet(input_args[0], input_args[1]) if additional else build_simple_resnet(input_args),
        "transfer_resnet": lambda: build_pretrained_resnet(input_args) if not additional else \
                                    build_pretrained_resnet_features(input_args[0], input_args[1]),
        "big_CNN": lambda: build_big_cnn(input_args[0], input_args[1]),
        "simple_CNN": lambda: build_simple_cnn(input_args, use_additional_inputs=additional)
    }

    # Obtener y devolver el modelo
    return model_map.get(model_name, lambda: None)()  # Devuelve `None` si el modelo no existe

def crop_raster(input_path, crop_size):
    with rasterio.open(input_path) as src:
        if crop_size > src.width or crop_size > src.height:
            raise ValueError(f"Crop size {crop_size}x{crop_size} exceeds raster dimensions {src.width}x{src.height}")
        
        # Define the crop window for the central crop_size x crop_size pixels
        left = (src.width - crop_size) // 2
        top = (src.height - crop_size) // 2
        window = Window(left, top, crop_size, crop_size)
        
        # Read the cropped window data
        cropped_data = src.read(1,window=window)
    return cropped_data


def rotate_image(image, max_angle_degrees=10):
    """
    Rotate an image by a random angle within ±max_angle_degrees.
    Args:
        image: Input image tensor (H, W, C).
        max_angle_degrees: Maximum rotation angle in degrees.
    Returns:
        Rotated image tensor.
    """
    # Convert degrees to radians
    max_angle_radians = max_angle_degrees * tf.constant(3.14159265 / 180, dtype=tf.float32)

    # Generate a random angle for rotation
    angle = tf.random.uniform([], -max_angle_radians, max_angle_radians)

    # Rotate the image using TensorFlow's affine transformations
    image_center = tf.cast(tf.shape(image)[:2], tf.float32) / 2.0
    rotation_matrix = tf.convert_to_tensor([
        [tf.cos(angle), -tf.sin(angle), image_center[0] - image_center[0] * tf.cos(angle) + image_center[1] * tf.sin(angle)],
        [tf.sin(angle), tf.cos(angle), image_center[1] - image_center[0] * tf.sin(angle) - image_center[1] * tf.cos(angle)],
        [0.0, 0.0, 1.0]
    ])

    # Apply rotation
    rotated_image = tf.keras.layers.experimental.preprocessing.RandomRotation(
        factor=max_angle_degrees / 180)(image)
    return rotated_image

def rotate_image(image, max_angle_degrees=10):
    """
    Rotate an image by a random angle within ±max_angle_degrees using pure TensorFlow.
    Args:
        image: Input image tensor (H, W, C).
        max_angle_degrees: Maximum rotation angle in degrees.
    Returns:
        Rotated image tensor.
    """
    # Convert degrees to radians
    max_angle_radians = max_angle_degrees * tf.constant(3.14159265 / 180, dtype=tf.float32)

    # Generate random rotation angle in radians
    angle = tf.random.uniform([], -max_angle_radians, max_angle_radians)

    # Get image dimensions
    height, width = tf.shape(image)[0], tf.shape(image)[1]

    # Rotation matrix components
    cos_angle = tf.cos(angle)
    sin_angle = tf.sin(angle)

    # Construct the flattened transform matrix
    transform = tf.convert_to_tensor([
        cos_angle, -sin_angle, 0.0,  # First row: a, b, c
        sin_angle, cos_angle,  0.0,  # Second row: d, e, f
        0.0, 0.0
    ], dtype=tf.float32)

    # Add batch dimension to image and transform
    transform = tf.expand_dims(transform, axis=0)  # Shape [1, 6]
    rotated_image = tf.raw_ops.ImageProjectiveTransformV3(
        images=tf.expand_dims(image, axis=0),  # Add batch dimension
        transforms=transform,  # Use the transform matrix
        output_shape=[height, width],
        interpolation="BILINEAR",
        fill_value=0.0  # Fill outside pixels with 0 (black)
    )

    return tf.squeeze(rotated_image, axis=0)  # Remove batch dimension


def augment_data_v0(image, label, additional_inp=None):
    """
    Apply data augmentation using TensorFlow image processing functions.
    Args:
        image: A tensor of shape (H, W, C) where the first 3 channels are RGB.
        label: The scalar temperature value.
    Returns:
        Augmented image and label.
    """
    # Define augmentation parameters
    brightness_delta = 0.1
    contrast_lower = 0.9
    contrast_upper = 1.1
    flip_prob = 0.5

    # Separate RGB and other variables
    lst_rgb = image[:, :, :3]
    other_variables = image[:, :, 3:]

    # Brightness and contrast adjustments (RGB only) 
    lst_rgb = tf.image.random_brightness(lst_rgb, max_delta=brightness_delta)
    lst_rgb = tf.image.random_contrast(lst_rgb, lower=contrast_lower, upper=contrast_upper)

    # Recombine RGB and other variables
    full_image = tf.concat([lst_rgb, other_variables], axis=-1)

    # Random flips
    if tf.random.uniform([]) < flip_prob:
        full_image = tf.image.flip_left_right(full_image)
    if tf.random.uniform([]) < flip_prob:
        full_image = tf.image.flip_up_down(full_image)

    # Random rotation
    full_image = rotate_image(full_image, max_angle_degrees=10)
    print("Inside augment_data - Image Shape:", image.shape)
    if additional_inp is not None:
        print("Inside augment_data - Additional Input Shape:", additional_inp.shape)
        return (full_image, additional_inp), label

    return full_image, label

def count_images(dataset, augment):
    """
    Count the total number of images in a dataset.
    Args:
        dataset: A TensorFlow dataset containing (image, label) pairs.
    Returns:
        Total number of images in the dataset.
    """
    total_images = 0
    if augment:
        for _, images, _ in dataset:  # Iterate over (image, label) pairs
            # Check if `images` is a list/tuple (multiple inputs to the model)
            if isinstance(images, (list, tuple)):
                # Add the number of samples in the first input (usually the primary data)
                total_images += images[0].shape[0]
            else:
                # Single input case
                total_images += images.shape[0]
    else:
        for _, (images, _) in dataset:  # Iterate over (image, label) pairs
            # Check if `images` is a list/tuple (multiple inputs to the model)
            if isinstance(images, (list, tuple)):
                # Add the number of samples in the first input (usually the primary data)
                total_images += images[0].shape[0]
            else:
                # Single input case
                total_images += images.shape[0]
        
    return total_images


def augment_data_v1(inputs_label, additional_inputs=False, is_outlier=False): # reference = (month, coords)
    if additional_inputs:
        image = inputs_label[0][0]
        additional_inp = inputs_label[0][1]
        label = inputs_label[1]
    else:
        image = inputs_label[0]
        label = inputs_label[1]
        
    # is_outlier = outlier_mask[(reference, label)] # True
    
    # Generate augmented images
    rotations = [
        image,  # 0°
        tf.image.rot90(image, k=1),  # 90°
        tf.image.rot90(image, k=2),  # 180°
        tf.image.rot90(image, k=3),  # 270°
    ]
    if is_outlier:
        rotations += [tf.identity(rot) for rot in rotations]  # Duplicate for outliers

    # Ensure all outputs have the same rank
    if additional_inputs:
        augmented_outputs = [
            ((rot, tf.convert_to_tensor(additional_inp)), tf.convert_to_tensor(label)) for rot in rotations
        ]
    else:
        augmented_outputs = [
            (rot, tf.convert_to_tensor(label)) for rot in rotations
        ]
    return augmented_outputs

def augment_data_v2(idx, inputs_label, reference, outlier_mask, additional_inputs=False, visualize = False): 

    if additional_inputs:
        image = tf.cast(inputs_label[0][0], dtype=tf.float32)
        additional_inp = tf.cast(inputs_label[0][1], dtype=tf.float32)
        label = tf.cast(inputs_label[1], dtype=tf.float32)
    else:
        image = tf.cast(inputs_label[0], dtype=tf.float32)
        label = tf.cast(inputs_label[1], dtype=tf.float32)
    
    reference = (tf.cast(reference[0], dtype=tf.float32), tf.cast(reference[1], dtype=tf.float32))
    reference_month = tf.gather(reference[0], idx)
    reference_coords = tf.gather(reference[1], idx)

    reference_month = tf.expand_dims(reference_month, axis=0)  # Convierte a [1]
    label = tf.expand_dims(label, axis=0)  # Convierte a [1]
    
    def create_outlier_key(month, coords, lbl):
        # Asegurar que las coordenadas son una tupla y los valores son flotantes
        python_key = (float(month.numpy()), tuple(map(float, coords.numpy())), float(lbl.numpy()))
        return python_key
    
    # Convertir `reference_coords` en una tupla de Python directamente
    def get_outlier_status(month, coords, lbl):
        # Crear la clave como tupla compatible con las del diccionario
        python_key = (float(month), tuple(map(float, coords)), float(lbl))
        return outlier_mask.get(python_key, False)
    
    # Usar `tf.py_function` para consultar el diccionario
    is_outlier = tf.py_function(
        func=lambda month, coords, lbl: get_outlier_status(month, coords, lbl),
        inp=[reference_month, reference_coords, label],
        Tout=tf.bool
    ) 

    # Generate augmented images
    base_rotations = [
        image,  # 0°
        tf.image.rot90(image, k=1),  # 90°
        tf.image.rot90(image, k=2),  # 180°
        tf.image.rot90(image, k=3),  # 270°
    ]
   
    rotations = tf.cond(
        is_outlier,
        lambda: base_rotations + base_rotations,  # if is_outlier duplicate rotations
        lambda: base_rotations + [tf.zeros_like(base_rotations[0]) for _ in base_rotations]  # if not is_outlier fill in with False
    )
    valid_mask = tf.logical_not(tf.reduce_all(tf.equal(rotations, 0), axis=[1, 2, 3]))
    rotations = tf.boolean_mask(rotations, valid_mask)

    def pack_output(rot):
        if additional_inputs:
            return ((rot, tf.convert_to_tensor(additional_inp)), label)
        else:
            return (rot, label)

    augmented_outputs = tf.map_fn(pack_output, rotations, fn_output_signature=(tf.TensorSpec(None, tf.float32), tf.TensorSpec(None, tf.float32)))
    
    if visualize:
        visualize_augmented_outputs = [
            {"image": rot, "label": label, "is_outlier": is_outlier} for rot in rotations
        ]
        return visualize_augmented_outputs

            
    return augmented_outputs

def augment_data(idx, inputs_label, reference, outlier_mask, W, additional_inputs=False, visualize=False): 
    if additional_inputs:
        image = tf.cast(inputs_label[0][0], dtype=tf.float32)
        additional_inp = tf.cast(inputs_label[0][1], dtype=tf.float32)
        label = tf.cast(inputs_label[1], dtype=tf.float32)
    else:
        image = tf.cast(inputs_label[0], dtype=tf.float32)
        label = tf.cast(inputs_label[1], dtype=tf.float32)
    
    reference = (tf.cast(reference[0], dtype=tf.float32), tf.cast(reference[1], dtype=tf.float32))
    reference_month = tf.gather(reference[0], idx)
    reference_coords = tf.gather(reference[1], idx)

    reference_month = tf.expand_dims(reference_month, axis=0)  # Convierte a [1]
    label = tf.expand_dims(label, axis=0) 

    def get_outlier_status(month, coords, lbl):
        python_key = (float(month), tuple(map(float, coords)), float(lbl))
        return outlier_mask.get(python_key, False)
    
    is_outlier = tf.py_function(
        func=lambda month, coords, lbl: get_outlier_status(month, coords, lbl),
        inp=[reference_month, reference_coords, label],
        Tout=tf.bool
    )

    # Generate augmented images
    base_rotations = [
        image,  # 0°
        tf.image.rot90(image, k=1),  # 90°
        tf.image.rot90(image, k=2),  # 180°
        tf.image.rot90(image, k=3),  # 270°
    ]

    rotations = tf.cond(
        is_outlier,
        lambda: base_rotations + base_rotations,  # 7 augmentations for outliers
        lambda: base_rotations + [tf.zeros_like(base_rotations[0]) for _ in base_rotations] # 4 augmentations for non-outliers
    )

    valid_mask = tf.logical_not(tf.reduce_all(tf.equal(rotations, 0), axis=[1, 2, 3]))
    rotations = tf.boolean_mask(rotations, valid_mask)
    num_channels = tf.get_static_value(tf.shape(image)[-1])
    rotations = tf.map_fn(
                    lambda rot: tf.ensure_shape(rot, [W, W, num_channels]),
                    rotations
                )


    def pack_output(rot):
        if additional_inputs:
            return idx, (rot, tf.convert_to_tensor(additional_inp)), label
        else:
            return idx, rot, label

    if visualize:
        visualize_augmented_outputs = [
            {"image": rot, "label": label, "is_outlier": is_outlier} for rot in rotations
        ]
        return visualize_augmented_outputs
        
    augmented_dataset = tf.data.Dataset.from_tensor_slices(rotations).map(
        lambda rot: pack_output(rot),
        num_parallel_calls=tf.data.AUTOTUNE
    )
    
    #augmented_dataset = augmented_dataset.concatenate(augmented_dataset).concatenate(augmented_dataset)

    return augmented_dataset



    
def load_set(data_folder, inputs, split_set, var_channels, var_position):
    
    set_dir = os.path.join(data_folder, split_set)
    input_data = np.load(os.path.join(set_dir, 'input_data.npy'))
    target = np.load(os.path.join(set_dir, 'target_data.npy'))
    additional_inputs = np.load(os.path.join(set_dir, 'additional_inputs.npy'))

    # Initialize inputs
    image_inputs = []  # To store image-based inputs
    vector_inputs = []  # To store vector-based inputs
    
    # Build the inputs based on the variables in the list `inputs`
    for inp in inputs:
        if inp in var_channels: # Add image channels
            channel = var_channels[inp] 
            image_inputs.append(input_data[..., channel:channel + 1])  # Slicing to preserve dimensions
            
        elif inp in var_position: # Add vector-based inputs from `additional_inputs`
            position = var_position[inp]
            if isinstance(position, list):
                vector_inputs.append(additional_inputs[:, position])
            else:
                vector_inputs.append(additional_inputs[:, position:position + 1])
                
    # Combine image inputs along the last axis
    image_inputs = np.concatenate(image_inputs, axis=-1) if image_inputs else None
    
    # Combine vector inputs along the last axis
    vector_inputs = np.concatenate(vector_inputs, axis=-1) if vector_inputs else None
    
    if vector_inputs is not None:  # If vector inputs are present
        input_args = (image_inputs.shape[1:], vector_inputs.shape[1])
        model_input = [image_inputs, vector_inputs]
        
    else:  # Only image inputs
        input_args = image_inputs.shape[1:]
        model_input = image_inputs
        
    if isinstance(model_input, list):
        print(f"{split_set} model input shapes: {[x.shape for x in model_input]}")
    else:
        print(f"{split_set} model input shape: {model_input.shape}")
        
    return model_input, additional_inputs, target #

def prepare_dataset_v0(ds, target, conditioned, reference=None, augment=False):
    inputs = (ds[0], ds[1]) if conditioned else ds
    dataset = tf.data.Dataset.from_tensor_slices((inputs, target))
    
    if augment:
        dataset_with_index = dataset.enumerate()
        with open('../data/external/outlier_dict.json', 'r') as json_file:
            loaded_dict = json.load(json_file)

        outlier_mask = {eval(key): value for key, value in loaded_dict.items()}

        dataset = dataset_with_index.map( \
            lambda idx, inputs_target: (augment_data(idx, inputs_target, reference, outlier_mask, \
                                                     additional_inputs = conditioned, visualize = False)), \
                                        num_parallel_calls=tf.data.AUTOTUNE)
    return dataset

def prepare_dataset(ds, target, conditioned, W=128, reference=None, augment=False):
    inputs = (ds[0], ds[1]) if conditioned else ds
    dataset = tf.data.Dataset.from_tensor_slices((inputs, target))
    
    dataset_with_indices = dataset.enumerate() 
    
    if augment:
        with open('../data/external/outlier_dict.json', 'r') as json_file:
            loaded_dict = json.load(json_file)
        outlier_mask = {eval(key): value for key, value in loaded_dict.items()}

        def augment_with_index(idx, inputs_target):
            augmented = augment_data(
                idx, inputs_target, reference, outlier_mask, W,
                additional_inputs=conditioned, visualize=False
            )
            return augmented

        dataset_with_indices = dataset_with_indices.flat_map(
            augment_with_index
        )
    
    return dataset_with_indices


def load_numpy_dataset(file_path):
    data = np.load(file_path + ".npz")
    indices = data["indices"]
    inputs = data["inputs"]
    additional_inputs = data['additional_inputs']
    print(inputs.shape,additional_inputs.shape)
    targets = data["targets"]
    return tf.data.Dataset.from_tensor_slices((indices, ((inputs,additional_inputs), targets)))


def select_channels(dataset, channels_to_keep):
    """
    Preprocess the dataset to keep only specified channels from the inputs.

    Args:
    - dataset: tf.data.Dataset with structure (index, ((image_inputs, additional_inputs), target)).
    - channels_to_keep: List of channel indices to keep (e.g., [0, 2]).

    Returns:
    - A preprocessed tf.data.Dataset.
    """
    def preprocess(index, inputs_targets):
        inputs, target = inputs_targets
        image_inputs, additional_inputs = inputs

        # Select specific channels
        selected_image_inputs = tf.gather(image_inputs, channels_to_keep, axis=-1)

        # Return the new structure
        return index, ((selected_image_inputs, additional_inputs), target)

    return dataset.map(preprocess)


def augment_data(idx, inputs_label, reference, outlier_mask, W, additional_inputs=False, visualize=False, augment_iterations=1): 
    """
    Augments the data by applying transformations to the input images. Allows for multiple augmentation iterations.

    Args:
        idx (int): Index of the data point.
        inputs_label (tuple): Tuple containing the image and label (and additional inputs if applicable).
        reference (tuple): Reference data containing months and coordinates.
        outlier_mask (dict): Dictionary mapping (month, coords, label) to outlier status.
        W (int): Image width/height.
        additional_inputs (bool): Whether additional inputs are included.
        visualize (bool): Whether to return augmented outputs for visualization.
        augment_iterations (int): Number of augmentations to apply per image.

    Returns:
        tf.data.Dataset: Dataset containing augmented images and labels.
    """
    if additional_inputs:
        image = tf.cast(inputs_label[0][0], dtype=tf.float32)
        additional_inp = tf.cast(inputs_label[0][1], dtype=tf.float32)
        label = tf.cast(inputs_label[1], dtype=tf.float32)
    else:
        image = tf.cast(inputs_label[0], dtype=tf.float32)
        label = tf.cast(inputs_label[1], dtype=tf.float32)
    
    reference = (tf.cast(reference[0], dtype=tf.float32), tf.cast(reference[1], dtype=tf.float32))
    reference_month = tf.gather(reference[0], idx)
    reference_coords = tf.gather(reference[1], idx)

    reference_month = tf.expand_dims(reference_month, axis=0)  # Convert to [1]
    label = tf.expand_dims(label, axis=0) 

    def get_outlier_status(month, coords, lbl):
        python_key = (float(month), tuple(map(float, coords)), float(lbl))
        return outlier_mask.get(python_key, False)
    
    is_outlier = tf.py_function(
        func=lambda month, coords, lbl: get_outlier_status(month, coords, lbl),
        inp=[reference_month, reference_coords, label],
        Tout=tf.bool
    )

    # Generate base rotations
    base_rotations = [
        image,  # 0°
        tf.image.rot90(image, k=1),  # 90°
        tf.image.rot90(image, k=2),  # 180°
        tf.image.rot90(image, k=3),  # 270°
    ]

    base_mirrors = [
        tf.image.flip_left_right(image),  # Mirror left-to-right
        tf.image.flip_up_down(image),    # Mirror top-to-bottom
    ]

    all_base_transformations = base_rotations + base_mirrors
    
    def augment(transformations, iterations):
        """Generates multiple augmentations by repeating the transformations."""
        return transformations * iterations

    # Apply augmentations based on outlier status
    transformations = tf.cond(
        is_outlier,
        lambda: augment(all_base_transformations, max(2, augment_iterations * 2)),
        lambda: augment(all_base_transformations, augment_iterations)
    )


    valid_mask = tf.logical_not(tf.reduce_all(tf.equal(transformations, 0), axis=[1, 2, 3]))
    transformations = tf.boolean_mask(transformations, valid_mask)
    num_channels = tf.get_static_value(tf.shape(image)[-1])
    rotations = tf.map_fn(
                    lambda trans: tf.ensure_shape(trans, [W, W, num_channels]),
                    transformations
                )

    def pack_output(trans):
        if additional_inputs:
            return idx, (trans, tf.convert_to_tensor(additional_inp)), label
        else:
            return idx, trans, label

    if visualize:
        visualize_augmented_outputs = [
            {"image": trans, "label": label, "is_outlier": is_outlier} for trans in transformations
        ]
        return visualize_augmented_outputs
        
    augmented_dataset = tf.data.Dataset.from_tensor_slices(transformations).map(
        lambda trans: pack_output(trans),
        num_parallel_calls=tf.data.AUTOTUNE
    )

    return augmented_dataset

def prepare_dataset(ds, target, conditioned, W=128, reference=None, augment=False, augment_iterations=1):
    """
    Prepares the dataset with optional augmentation.

    Args:
        ds (tuple): Tuple of input data.
        target (array): Target data.
        conditioned (bool): Whether additional inputs are included.
        W (int): Image width/height.
        reference (tuple): Reference data containing months and coordinates.
        augment (bool): Whether to apply augmentation.
        augment_iterations (int): Number of augmentations to apply per image.

    Returns:
        tf.data.Dataset: Prepared dataset.
    """
    inputs = (ds[0], ds[1]) if conditioned else ds
    dataset = tf.data.Dataset.from_tensor_slices((inputs, target))
    
    dataset_with_indices = dataset.enumerate() 
    
    if augment:
        with open('../data/external/outlier_dict.json', 'r') as json_file:
            loaded_dict = json.load(json_file)
        outlier_mask = {eval(key): value for key, value in loaded_dict.items()}

        def augment_with_index(idx, inputs_target):
            augmented = augment_data(
                idx, inputs_target, reference, outlier_mask, W,
                additional_inputs=conditioned, visualize=False, augment_iterations=augment_iterations
            )
            return augmented

        dataset_with_indices = dataset_with_indices.flat_map(
            augment_with_index
        )
    
    return dataset_with_indices

from keras import backend as K

# Liberar memoria de la GPU después de cada experimento
def clear_memory():
    K.clear_session()
    tf.compat.v1.reset_default_graph()
    # Opcionalmente, también puedes usar:
    tf.keras.backend.clear_session() 
