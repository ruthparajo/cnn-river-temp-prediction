import rasterio
import pandas as pd
import os
import re
import subprocess
import numpy as np
from skimage.transform import resize
import openpyxl
from openpyxl import load_workbook
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from rasterio.transform import from_origin
import shutil
import matplotlib.pyplot as plt


def load_raster(filepath,rgb = True):
  with rasterio.open(filepath) as src:
    if rgb:
      red = src.read(1)
      green = src.read(2)
      blue = src.read(3)
      rgb = np.dstack((red, green, blue))
      return rgb, src
    else:
      return src.read(1), src
      
def resize_image(image, new_width, new_height):
    return resize(image, (new_height, new_width), preserve_range=True)

def extract_year_month_from_filename(filename):
    # Function to extract the year and month from the filename
    # Adjust regex to match the YYYY-MM format
    match = re.search(r'(\d{4})-(\d{2})', filename)
    if match:
        return match.group(0)  # Return the matched year and month as a string (YYYY-MM)
    return None

def load_data(dir_paths, W=None, rgb=[], show=None):
    data = {}
    time_slots = []
    for i, path in enumerate(dir_paths):
        image_files = os.listdir(path)

        # Filter out files without a valid date in the filename
        valid_image_files = [img for img in image_files if extract_year_month_from_filename(img) is not None]

        # Sort the valid image files by year and month extracted from the filename
        valid_image_files = sorted(valid_image_files, key=lambda x: extract_year_month_from_filename(x))
        
        times = []
        imgs_data = []
        for img_file in valid_image_files:
            imagen_path = os.path.join(path, img_file)
            img_data, meta = load_raster(imagen_path, rgb[i])  # Load raster data

            if W:  # Resize the image if W is provided
                img_data = resize_image(img_data, W, W)

            imgs_data.append(img_data)
            times.append(extract_year_month_from_filename(img_file))
            
        data[path] = np.array(imgs_data)
        time_slots.append(times)

    return data, time_slots

def split_data(X, y, validation_split=0.1, test_split=0.1):
    """
    Splits the data into training, validation, and test sets.

    Parameters:
    X (np.array): Input data (e.g., images).
    y (np.array): Target data (e.g., temperature values).
    validation_split (float): Proportion of data to use for validation set.
    test_split (float): Proportion of data to use for test set.

    Returns:
    train_input (np.array): Training data inputs normalized.
    train_target (np.array): Training data targets.
    validation_input (np.array): Validation data inputs normalized.
    validation_target (np.array): Validation data targets.
    test_input (np.array): Test data inputs normalized.
    test_target (np.array): Test data targets.
    """

    n = X.shape[0]  # Number of samples
    rem_index = np.arange(n)  # Array of all indices

    # Select validation indices
    validation_index = np.random.choice(rem_index, int(validation_split * n), replace=False)
    rem_index = list(set(rem_index) - set(validation_index))

    # Select test indices from remaining
    test_index = np.random.choice(rem_index, int(test_split * n), replace=False)
    train_index = list(set(rem_index) - set(test_index))  # Remaining for training

    # Split data into train, validation, and test sets
    validation_input = X[validation_index, :] / 255.0  # Normalize inputs
    validation_target = y[validation_index, :]

    test_input = X[test_index, :] / 255.0  # Normalize inputs
    test_target = y[test_index, :]

    train_input = X[train_index, :] / 255.0  # Normalize inputs
    train_target = y[train_index, :]

    return train_input, train_target, validation_input, validation_target, test_input, test_target

def evaluate_model(y_true, y_pred):
    """
    Returns useful metrics for evaluating a linear regression model for temperature prediction.

    Parameters:
    y_true: array-like, true temperature values.
    y_pred: array-like, predicted temperature values.

    Returns:
    dict containing all useful metrics.
    """

    # Mean Absolute Error (MAE)
    mae = mean_absolute_error(y_true, y_pred)

    # Mean Squared Error (MSE)
    mse = mean_squared_error(y_true, y_pred)

    # Root Mean Squared Error (RMSE)
    rmse = np.sqrt(mse)

    # R-squared (R²)
    r2 = r2_score(y_true, y_pred)

    # Mean Absolute Percentage Error (MAPE)
    epsilon = 1e-10  # Small value to avoid division by zero
    mape = np.mean(np.abs((y_true - y_pred) / (y_true + epsilon))) * 100

    # Mean Squared Error (MSE) sample-wise
    sample_wise_mse = []
    # Loop through each sample and compute the MSE for that sample
    for i in range(y_true.shape[0]):
        # Flatten the true and predicted values for this sample
        y_true_flatten = y_true[i].flatten()
        y_pred_flatten = y_pred[i].flatten()
        # Calculate MSE for this sample
        mse = mean_squared_error(y_true_flatten, y_pred_flatten)
        sample_wise_mse.append(mse)
    
    average_mse = np.mean(sample_wise_mse)

    # Create a dictionary with the metrics
    metrics = {
        'MAE': mae,
        'MSE': mse,
        'RMSE': rmse,
        'R²': r2,
        'MAPE (%)': mape,
        'MSE sample-wise': average_mse
    }

    return metrics

def update_excel_with_results(file_path, model_name, metrics, details):
    # Check if the Excel file exists
    if not os.path.exists(file_path):
        # Create a new Excel workbook and add a sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Model Results"
        # Write the header based on the metric keys
        header = ["Model Name"] + list(metrics.keys()) + ["Details"]
        sheet.append(header)
    else:
        # Load existing Excel workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active

    # Find the next empty row
    next_row = sheet.max_row + 1
    
    # Prepare the row with model name, metrics, and details
    row = [model_name] + list(metrics.values()) + [details]
    
    # Append the row data to the next available row
    for col_num, value in enumerate(row, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)
    
    # Save the Excel file
    workbook.save(file_path)

def save_excel(file_path, dicc, excel=None):
    # Check if the Excel file exists
    if not os.path.exists(file_path):
        # Create a new Excel workbook and add a sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        if excel == 'Results':
            sheet.title = "Model Results"
        else:
            sheet.title = "Model Details"
        # Write the header based on the metric keys
        header = list(dicc.keys())
        sheet.append(header)
    else:
        # Load existing Excel workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active

    # Find the next empty row
    next_row = sheet.max_row + 1
    
    # Prepare the row with model name, metrics, and details
    row = list(dicc.values())
    
    # Append the row data to the next available row
    for col_num, value in enumerate(row, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)
    
    # Save the Excel file
    workbook.save(file_path)

def split_data_df(X, y, validation_split=0.1, test_split=0.1):
    n = X.shape[0]  # Número de muestras
    rem_index = np.arange(n)  # Array de todos los índices

    # Seleccionar índices de validación
    validation_index = np.random.choice(rem_index, int(validation_split * n), replace=False)
    rem_index = list(set(rem_index) - set(validation_index))

    # Seleccionar índices de prueba de los restantes
    test_index = np.random.choice(rem_index, int(test_split * n), replace=False)
    train_index = list(set(rem_index) - set(test_index))  # El resto para entrenamiento

    # Dividir los datos en conjuntos de entrenamiento, validación y prueba
    validation_input = X.iloc[validation_index, :].values / 255.0  # Normalizar inputs si es necesario
    validation_target = y.iloc[validation_index].values

    test_input = X.iloc[test_index, :].values / 255.0  # Normalizar inputs
    test_target = y.iloc[test_index].values

    train_input = X.iloc[train_index, :].values / 255.0  # Normalizar inputs
    train_target = y.iloc[train_index].values

    return train_input, train_target, validation_input, validation_target, test_input, test_target

def plot_image(image, factor=1.0, clip_range=None, **kwargs):
    """
    Utility function for plotting RGB images.
    """
    fig, ax = plt.subplots(nrows=1, ncols=1, figsize=(15, 15))
    if clip_range is not None:
        ax.imshow(np.clip(image * factor, *clip_range), **kwargs)
    else:
        ax.imshow(image * factor, **kwargs)
    ax.set_xticks([])
    ax.set_yticks([])

def clear_directory(directory):
    # Function to clear the contents of a directory
    if os.path.exists(directory):
        for file in os.listdir(directory):
            file_path = os.path.join(directory, file)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.remove(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print(f'Error deleting {file_path}. Reason: {e}')

def save_raster(raster_array,filepath,shp):
    resolution = 30
    x_min, y_min, x_max, y_max = shp.total_bounds
    transform = from_origin(x_min, y_max, resolution, resolution)

    with rasterio.open(
        filepath,
        'w',
        driver='GTiff',
        height=raster_array.shape[0],
        width=raster_array.shape[1],
        count=len(raster_array.shape),
        dtype=raster_array.dtype,
        crs=shp.crs.to_string(),  # Ensure correct CRS
        transform=transform,
        nodata=0.0
    ) as dst:
        if len(raster_array.shape) == 3:
            dst.write(raster_array[:, :, 0], 1)  # Red channel
            dst.write(raster_array[:, :, 1], 2)  # Green channel
            dst.write(raster_array[:, :, 2], 3)  # Blue channel
        else:
            dst.write(raster_array, 1)

def distance_matrix(x0, y0, x1, y1):
    """
    Calculate distance matrix.
    Note: from <http://stackoverflow.com/questions/1871536>
    """

    obs = np.vstack((x0, y0)).T
    interp = np.vstack((x1, y1)).T

    d0 = np.subtract.outer(obs[:, 0], interp[:, 0])
    d1 = np.subtract.outer(obs[:, 1], interp[:, 1])

    # calculate hypotenuse
    return np.hypot(d0, d1)

def simple_idw(x, y, z, xi, yi, beta=2):
    """
    Simple inverse distance weighted (IDW) interpolation
    x`, `y`,`z` = known data arrays containing coordinates and data used for interpolation
    `xi`, `yi` =  two arrays of grid coordinates
    `beta` = determines the degree to which the nearer point(s) are preferred over more distant points.
            Typically 1 or 2 (inverse or inverse squared relationship)
    """
    print('Fent')

    dist = distance_matrix(x, y, xi, yi)

    
    '''weights = 1.0 / (dist + 1e-12) ** beta
    
    # Verificar si la suma de pesos es válida
    sum_weights = weights.sum(axis=0)
    
    # Normalizar los pesos
    weights /= sum_weights + 1e-12'''

    # In IDW, weights are 1 / distance
    # weights = 1.0/(dist+1e-12)**power
    weights = dist ** (-beta)
    
    # Make weights sum to one
    weights /= weights.sum(axis=0)

    weights = np.nan_to_num(weights, nan=0.0)
    z = np.nan_to_num(z, nan=0.0)

    # Multiply the weights for each interpolated point by all observed Z-values
    return np.dot(weights.T, z)

def project_linestrings_to_points(gdf):
    x_coords = []
    y_coords = []
    z_coords = []

    # Extract coordinates from LineStrings and store them as points
    for i, row in gdf.iterrows():
        geom = row.geometry
        if geom.geom_type == 'LineString':
            for coord in row.geometry.coords:
                x_coords.append(coord[0])
                y_coords.append(coord[1])
                if len(coord) == 3:
                    z_coords.append(coord[2])

        elif geom.geom_type == 'MultiLineString':
            # Si es MultiLineString, recorre cada LineString
            for linestring in geom.geoms:
                for coord in linestring.coords:
                    x_coords.append(coord[0])
                    y_coords.append(coord[1])
                    if len(coord) == 3:
                        z_coords.append(coord[2])


    # Plot the points on a 2D grid
    plt.figure(figsize=(10, 10))
    plt.scatter(x_coords, y_coords, color='blue', s=10, label='Coordinates')

    # Customize the plot
    plt.title('Projected Coordinates (Points) in EPSG:2056', fontsize=15)
    plt.xlabel('Easting (meters)', fontsize=12)
    plt.ylabel('Northing (meters)', fontsize=12)

    # Add gridlines
    plt.grid(True)

    # Show the plot
    plt.legend()
    plt.show()
    return x_coords, y_coords, z_coords